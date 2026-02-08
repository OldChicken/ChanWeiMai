from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


def build_template(output_path: Path, repo_root: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "商品利润率明细"

    # Collect store list from repo if available
    stores = []
    store_dir = repo_root / "店铺档案"
    if store_dir.exists():
        for p in store_dir.iterdir():
            if p.is_dir():
                stores.append(p.name)
    if not stores:
        stores = ["杭州-首店"]

    # Dictionary sheet for data validation lists
    dict_ws = wb.create_sheet("字典")
    dict_ws["A1"].value = "门店列表"
    for idx, name in enumerate(stores, start=2):
        dict_ws.cell(row=idx, column=1, value=name)

    dict_ws["B1"].value = "类别列表"
    categories = ["主食", "小吃", "饮品", "其他"]
    for idx, name in enumerate(categories, start=2):
        dict_ws.cell(row=idx, column=2, value=name)

    # Headers
    headers = [
        "日期", "门店", "商品名称", "SKU", "类别", "供应商",
        "进货价", "成本税率", "运/包/损", "总成本",
        "建议售价", "实际售价", "折扣金额",
        "平台佣金率", "平台佣金",
        "支付手续费率", "支付手续费",
        "包装费", "配送费", "运营分摊",
        "销项税率", "税金",
        "毛利", "毛利率",
        "净利", "净利率",
        "备注",
    ]

    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFBBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Column widths
    widths = {
        1: 12, 2: 16, 3: 20, 4: 12, 5: 12, 6: 16,
        7: 12, 8: 12, 9: 12, 10: 12,
        11: 12, 12: 12, 13: 12,
        14: 12, 15: 12,
        16: 14, 17: 14,
        18: 10, 19: 10, 20: 12,
        21: 12, 22: 12,
        23: 12, 24: 10,
        25: 12, 26: 10,
        27: 20,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Freeze header and add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data validation for 门店 (B) and 类别 (E)
    store_last_row = 1 + len(stores)
    cat_last_row = 1 + len(categories)
    dv_store = DataValidation(type="list", formula1=f"=字典!$A$2:$A${store_last_row}", allow_blank=True)
    dv_cat = DataValidation(type="list", formula1=f"=字典!$B$2:$B${cat_last_row}", allow_blank=True)
    ws.add_data_validation(dv_store)
    ws.add_data_validation(dv_cat)
    dv_store.add(f"B2:B1001")
    dv_cat.add(f"E2:E1001")

    # Formats
    currency_fmt = '"¥"#,##0.00'
    percent_fmt = "0.00%"
    date_fmt = "yyyy-mm-dd"

    # Apply formats to columns for rows 2..1001
    for r in range(2, 1002):
        ws.cell(row=r, column=1).number_format = date_fmt  # 日期
        for c in (7, 9, 10, 11, 12, 13, 15, 17, 18, 19, 20, 22, 23, 25):
            ws.cell(row=r, column=c).number_format = currency_fmt
        for c in (8, 14, 16, 21, 24, 26):
            ws.cell(row=r, column=c).number_format = percent_fmt

        # Formulas
        # J: 总成本 = G*(1+H)+I
        ws.cell(row=r, column=10, value=f"=G{r}*(1+H{r})+I{r}")
        # O: 平台佣金 = (L-M)*N
        ws.cell(row=r, column=15, value=f"=(L{r}-M{r})*N{r}")
        # Q: 支付手续费 = (L-M)*P
        ws.cell(row=r, column=17, value=f"=(L{r}-M{r})*P{r}")
        # V: 税金 = (L-M)*U
        ws.cell(row=r, column=22, value=f"=(L{r}-M{r})*U{r}")
        # W: 毛利 = (L-M) - O - Q - R - S - T - J
        ws.cell(row=r, column=23, value=f"=(L{r}-M{r})-O{r}-Q{r}-R{r}-S{r}-T{r}-J{r}")
        # X: 毛利率 = IFERROR(W/(L-M), "")
        ws.cell(row=r, column=24, value=f"=IFERROR(W{r}/(L{r}-M{r}),"")")
        # Y: 净利 = W - V
        ws.cell(row=r, column=25, value=f"=W{r}-V{r}")
        # Z: 净利率 = IFERROR(Y/(L-M), "")
        ws.cell(row=r, column=26, value=f"=IFERROR(Y{r}/(L{r}-M{r}),"")")

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    repo_root = Path(__file__).resolve().parent.parent
    out_file = repo_root / "模板与示例" / "财务模板" / "商品利润率模板.xlsx"
    build_template(out_file, repo_root)
    print(f"模板已生成: {out_file}")
